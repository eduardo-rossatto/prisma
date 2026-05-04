import re
import zipfile
import streamlit as st
import openpyxl
import streamlit.components.v1 as components
from io import BytesIO
from datetime import date, datetime
from openpyxl.styles import PatternFill, Font

st.set_page_config(
    page_title="PRISMA",
    layout="wide",
    page_icon="🤝",
    initial_sidebar_state="expanded",
)

TEMPLATE_PATH = "template/Prisma - Template.xlsx"

GESTORES = {
    "Amanda Cristine Gonçalves": {
        "email": "amanda.goncalves@biotrop.com.br",
        "tel": "(41) 3099-7300 | (19) 3886-4140 | (41) 99143-1823",
    },
    "Tatiane Carvalho": {
        "email": "tatiane.carvalho@biotrop.com.br",
        "tel": "(41) 3099-7300 | (19) 3886-4140 | (41) 99293-7173",
    },
}

DOCS_LTDA = [
    "Documentos pessoais dos sócios e cônjuges (RG / CPF)",
    "Comprovante de endereço (sócios)",
    "Certidão de casamento dos sócios (se aplicável)",
    "Contrato social e última alteração contratual",
    "Balanço Patrimonial e DRE — últimos 2 anos (assinados)",
    "Imposto de Renda dos sócios — último exercício",
]

DOCS_COOP = [
    "Documentos pessoais dos dirigentes",
    "Comprovante de endereço dos dirigentes",
    "Certidão de casamento dos sócios (se aplicável)",
    "Estatuto Social",
    "Ata de Eleição da Diretoria vigente",
    "Balanço Patrimonial e DRE — últimos 2 anos (assinados)",
]

# ══════════════════════════════════════════════════════════════════════════════
# CSS
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600&family=Space+Grotesk:wght@500;600;700&family=IBM+Plex+Mono:wght@400;500&display=swap');

html, body, [class*="css"] { font-family: 'Inter', sans-serif !important; }
#MainMenu, footer, header, [data-testid="stDecoration"], [data-testid="stToolbar"] { display: none !important; }

.stApp,
[data-testid="stAppViewContainer"],
[data-testid="stMain"],
[data-testid="stMainBlockContainer"],
.main .block-container { background-color: #0f0f0e !important; }

[data-testid="stSidebar"] {
    background-color: #141312 !important;
    border-right: 1px solid rgba(255,255,255,0.06) !important;
}
[data-testid="stSidebar"] section { padding: 0 20px !important; }

p, span, div, li, a, [data-testid="stMarkdownContainer"] p { color: #f8f8f7 !important; }
h1, h2, h3, h4 { color: #f8f8f7 !important; font-family: 'Space Grotesk', sans-serif !important; }

label,
[data-testid="stWidgetLabel"] > div,
[data-testid="stWidgetLabel"] p,
[data-testid="stTextInput"] label,
[data-testid="stSelectbox"] label,
[data-testid="stTextArea"] label,
[data-testid="stDateInput"] label,
[data-testid="stNumberInput"] label {
    color: #766f6b !important;
    font-size: 0.68rem !important;
    font-weight: 500 !important;
    letter-spacing: 0.08em !important;
    text-transform: uppercase !important;
}

input[type="text"], input[type="number"], input[type="email"],
input[type="date"], input[type="time"], textarea {
    background-color: #1e1d1b !important;
    border: 1px solid rgba(255,255,255,0.09) !important;
    border-radius: 8px !important;
    color: #f8f8f7 !important;
    font-size: 0.86rem !important;
    font-family: 'Inter', sans-serif !important;
    transition: border-color .15s ease !important;
}
input:focus, textarea:focus {
    border-color: rgba(244,62,1,.45) !important;
    box-shadow: 0 0 0 3px rgba(244,62,1,.08) !important;
    outline: none !important;
}

[data-testid="stTextInput"]:has(input[aria-label*="CNPJ"]) input,
[data-testid="stTextInput"]:has(input[aria-label*="CEP"]) input,
[data-testid="stTextInput"]:has(input[aria-label*="CNAE"]) input {
    font-family: 'IBM Plex Mono', monospace !important;
    letter-spacing: .05em !important;
}

[data-baseweb="select"] > div {
    background-color: #1e1d1b !important;
    border-color: rgba(255,255,255,0.09) !important;
    border-radius: 8px !important;
    color: #f8f8f7 !important;
}
[data-baseweb="select"] [data-testid="stMarkdownContainer"] p,
[data-baseweb="select"] span { color: #f8f8f7 !important; font-size: .86rem !important; }
[data-baseweb="popover"],
[data-baseweb="popover"] ul,
[data-baseweb="popover"] [role="listbox"] {
    background-color: #201f1d !important;
    border: 1px solid rgba(255,255,255,0.1) !important;
    border-radius: 10px !important;
}
[data-baseweb="popover"] [role="option"] { color: #f8f8f7 !important; }
[data-baseweb="popover"] [role="option"]:hover { background-color: rgba(244,62,1,.12) !important; }
[data-baseweb="popover"] [aria-selected="true"] { background-color: rgba(244,62,1,.18) !important; }

[data-testid="stDateInputField"] input { color: #f8f8f7 !important; }

/* ── file uploader ── */
[data-testid="stFileUploader"] section {
    background-color: #1e1d1b !important;
    border: 1px dashed rgba(255,255,255,0.12) !important;
    border-radius: 8px !important;
    padding: 8px 12px !important;
}
[data-testid="stFileUploader"] section:hover {
    border-color: rgba(244,62,1,.4) !important;
}
[data-testid="stFileUploaderDropzoneInstructions"] span { color: #766f6b !important; font-size: .72rem !important; }
[data-testid="stFileUploaderDropzoneInstructions"] small { color: #5a5450 !important; font-size: .62rem !important; }
[data-testid="stFileUploader"] button { color: #f8f8f7 !important; font-size: .7rem !important; }
[data-testid="stFileUploaderFileName"] { color: #10e68d !important; font-size: .75rem !important; }

.stTabs [data-baseweb="tab-list"] {
    background: transparent !important;
    border-bottom: 1px solid rgba(255,255,255,0.07) !important;
    gap: 0 !important; padding: 0 !important; margin-bottom: 28px !important;
}
.stTabs [data-baseweb="tab"] {
    background: transparent !important; color: #5a5450 !important;
    font-size: 0.75rem !important; font-weight: 500 !important;
    letter-spacing: .06em !important; text-transform: uppercase !important;
    border-radius: 0 !important; border: none !important;
    border-bottom: 2px solid transparent !important;
    padding: 10px 18px !important; margin-bottom: -1px !important;
    transition: color .15s ease !important;
}
.stTabs [aria-selected="true"] { color: #f8f8f7 !important; border-bottom-color: #f43e01 !important; }
.stTabs [data-baseweb="tab"]:hover { color: #a09890 !important; background: transparent !important; }
.stTabs [data-baseweb="tab-highlight"],
.stTabs [data-baseweb="tab-border"] { display: none !important; }
.stTabs [data-baseweb="tab-panel"] { padding: 0 !important; }

.stButton > button {
    background: transparent !important; border: 1px solid rgba(255,255,255,0.12) !important;
    color: #f8f8f7 !important; border-radius: 9999px !important;
    font-size: 0.75rem !important; font-weight: 500 !important;
    letter-spacing: .08em !important; text-transform: uppercase !important;
    padding: 8px 20px !important; transition: all .15s ease !important;
}
.stButton > button:hover { border-color: rgba(255,255,255,.3) !important; background: rgba(255,255,255,.04) !important; }
.stButton > button[kind="primary"] { background: #f43e01 !important; border-color: #f43e01 !important; color: #fff !important; font-weight: 600 !important; }
.stButton > button[kind="primary"]:hover { background: #d93500 !important; border-color: #d93500 !important; }
.stDownloadButton > button {
    background: rgba(16,230,141,.08) !important; border: 1px solid rgba(16,230,141,.25) !important;
    color: #10e68d !important; border-radius: 9999px !important;
    font-size: .75rem !important; font-weight: 600 !important;
    letter-spacing: .08em !important; text-transform: uppercase !important;
    width: 100% !important; transition: all .15s ease !important;
}
.stDownloadButton > button:hover { background: rgba(16,230,141,.14) !important; border-color: rgba(16,230,141,.45) !important; }

hr { border-color: rgba(255,255,255,0.06) !important; margin: 16px 0 !important; }
[data-testid="stAlert"] { background: rgba(244,62,1,.08) !important; border: 1px solid rgba(244,62,1,.2) !important; border-radius: 8px !important; color: #f8f8f7 !important; }
::-webkit-scrollbar { width: 4px; }
::-webkit-scrollbar-track { background: transparent; }
::-webkit-scrollbar-thumb { background: rgba(255,255,255,.1); border-radius: 4px; }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def fmt_date(d):
    if not d:
        return ""
    if isinstance(d, (date, datetime)):
        return d.strftime("%d/%m/%Y")
    return str(d)

def _safe_filename(name: str, max_len: int = 60) -> str:
    return re.sub(r'[\\/:*?"<>|]', '-', name).strip()[:max_len]

def fill_template(text_data: dict, file_data: dict = None) -> tuple:
    """
    Returns (excel_bytes, files_to_include).
    - Empty text values  → '-' + light-gray cell fill
    - None file entries  → 'Pendente' + light-gray cell fill
    - File dict entries  → =HYPERLINK(...) formula + blue underline font
    """
    if file_data is None:
        file_data = {}

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    link_font  = Font(color="0563C1", underline="single")

    files_to_include = {}

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not (cell.value and isinstance(cell.value, str)):
                    continue

                new_val    = cell.value
                has_empty  = False
                has_link   = False

                for key, value in text_data.items():
                    token = "{{" + key + "}}"
                    if token in new_val:
                        repl = str(value) if value else "-"
                        new_val = new_val.replace(token, repl)
                        if not value:
                            has_empty = True

                for key, finfo in file_data.items():
                    token = "{{" + key + "}}"
                    if token in new_val:
                        if finfo:
                            fname   = finfo["saved_name"]
                            display = finfo.get("display", fname)
                            fe = fname.replace('"', '""')
                            de = display.replace('"', '""')
                            new_val = new_val.replace(
                                token,
                                f'=HYPERLINK("documentos/{fe}","{de}")'
                            )
                            files_to_include[f"documentos/{fname}"] = finfo["bytes"]
                            has_link = True
                        else:
                            new_val = new_val.replace(token, "Pendente")
                            has_empty = True

                cell.value = new_val
                if has_link:
                    cell.font = link_font
                elif has_empty:
                    cell.fill = gray_fill

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), files_to_include


def build_zip(excel_bytes: bytes, files: dict, razao: str) -> bytes:
    zip_buf = BytesIO()
    safe = _safe_filename(razao)
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"PRISMA_{safe}.xlsx", excel_bytes)
        for path, content in files.items():
            zf.writestr(path, content)
    zip_buf.seek(0)
    return zip_buf.getvalue()


def sec(num_or_label, title, desc=""):
    st.markdown(f"""
    <div style="margin:4px 0 28px">
      <div style="font-size:.6rem;font-weight:600;letter-spacing:.22em;text-transform:uppercase;
                  color:#766f6b;margin-bottom:10px">{num_or_label}</div>
      <div style="font-family:'Space Grotesk',sans-serif;font-size:1.35rem;font-weight:600;
                  color:#f8f8f7;letter-spacing:-.015em;line-height:1.2">{title}</div>
      {"" if not desc else f'<div style="font-size:.8rem;color:#766f6b;margin-top:6px;line-height:1.6">{desc}</div>'}
      <div style="width:100%;height:1px;background:rgba(255,255,255,.06);margin-top:20px"></div>
    </div>
    """, unsafe_allow_html=True)

def sub(label):
    st.markdown(f"""
    <div style="font-size:.6rem;font-weight:600;letter-spacing:.18em;text-transform:uppercase;
                color:#5a5450;margin:22px 0 10px;padding-left:1px">{label}</div>
    """, unsafe_allow_html=True)

def col_head(*labels, cols):
    for col, lbl in zip(cols, labels):
        col.markdown(
            f"<div style='font-size:.6rem;font-weight:600;letter-spacing:.14em;"
            f"text-transform:uppercase;color:#5a5450;padding-bottom:6px'>{lbl}</div>",
            unsafe_allow_html=True,
        )

def sim_nao(label, key, lv="visible"):
    return st.selectbox(label, ["", "Sim", "Não"], key=key, label_visibility=lv)

def nivel(label, key, lv="visible"):
    return st.selectbox(label, ["", "1 — Baixo", "2 — Médio", "3 — Alto"], key=key, label_visibility=lv)

def s_doc(label, key, lv="visible"):
    return st.selectbox(label, ["", "Pendente", "Entregue", "Não se aplica"], key=key, label_visibility=lv)

def s_lic(label, key, lv="visible"):
    return st.selectbox(label, ["", "Válida", "Vencida", "Pendente", "Não se aplica"], key=key, label_visibility=lv)

def autofill_box(label, value):
    lbl_s = "color:#766f6b;font-size:.68rem;font-weight:500;letter-spacing:.08em;text-transform:uppercase"
    val_s = "color:#f8f8f7;font-size:.86rem" if value else "color:#3a3430;font-size:.86rem"
    st.markdown(
        f"<div style='margin-bottom:16px'>"
        f"<div style='{lbl_s};margin-bottom:4px'>{label}</div>"
        f"<div style='background:#1e1d1b;border:1px solid rgba(255,255,255,.06);border-radius:8px;"
        f"padding:9px 12px;min-height:38px'>"
        f"<span style='{val_s}'>{value or '—'}</span>"
        f"</div></div>",
        unsafe_allow_html=True,
    )

# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("""
    <div style="padding:32px 0 24px">
      <div style="font-family:'Space Grotesk',sans-serif;font-size:1.5rem;font-weight:700;
                  letter-spacing:.12em;color:#f8f8f7">PRISMA</div>
      <div style="font-size:.58rem;color:#5a5450;letter-spacing:.14em;
                  text-transform:uppercase;margin-top:5px">Ficha Técnica do Cliente</div>
    </div>
    """, unsafe_allow_html=True)
    st.divider()
    for num, label in [
        ("01","Cadastro"),("02","Endereços"),("03","Contatos"),
        ("04","Comercial"),("05","Regras"),("06","Licenças"),
        ("07","Estratégico"),("08","Crédito"),("09","Produtos"),
    ]:
        st.markdown(f"""
        <div style="display:flex;align-items:center;gap:12px;padding:6px 0">
          <span style="font-family:'IBM Plex Mono',monospace;font-size:.58rem;color:#3a3430;font-weight:500">{num}</span>
          <span style="font-size:.78rem;color:#766f6b;font-weight:400">{label}</span>
        </div>""", unsafe_allow_html=True)
    st.markdown("""<div style="position:fixed;bottom:28px;font-size:.6rem;color:#3a3430;
                letter-spacing:.06em">BIOTROP · uso interno</div>""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# TABS
# ══════════════════════════════════════════════════════════════════════════════
if "_click_tab" in st.session_state:
    idx = st.session_state.pop("_click_tab")
    components.html(f"""<script>
    setTimeout(function(){{
        var t=window.parent.document.querySelectorAll('[data-baseweb="tab"]');
        if(t[{idx}])t[{idx}].click();
    }},120);</script>""", height=0)

st.markdown("""
<div style="padding:28px 0 20px">
  <div style="font-family:'Space Grotesk',sans-serif;font-size:2rem;font-weight:700;
              color:#f8f8f7;letter-spacing:.08em;line-height:1">PRISMA</div>
  <div style="font-size:.72rem;color:#5a5450;margin-top:6px;letter-spacing:.04em">
    Partner · Relationship · Intelligence · Strategy · Mapping · Assessment
    &nbsp;·&nbsp; Ficha Técnica de Clientes
  </div>
</div>""", unsafe_allow_html=True)

tabs = st.tabs(["Cadastro","Endereços","Contatos","Comercial",
                "Regras","Licenças","Estratégico","Crédito","Produtos"])

# ─────────────────────────────────────────────
# 01 · CADASTRO
# ─────────────────────────────────────────────
with tabs[0]:
    sec("Bloco 01","Dados Cadastrais","Identificação legal e fiscal do cliente.")
    c1,c2,c3 = st.columns([3,3,2])
    with c1:
        st.text_input("Razão Social", key="razao_social")
        st.text_input("CNPJ", placeholder="00.000.000/0000-00", key="cnpj")
    with c2:
        st.text_input("Nome Fantasia", key="nome_fantasia")
        st.text_input("CNAE", placeholder="0000-0/00", key="cnae")
    with c3:
        st.text_input("Inscrição Estadual", key="inscricao_estadual")
        st.text_input("Inscrição Municipal", key="inscricao_municipal")
    c4,c5 = st.columns([2,6])
    with c4:
        st.date_input("Data de Abertura", value=None, key="data_abertura", format="DD/MM/YYYY")

    sub("Responsável pela Gestão de Carteira")
    gestor = st.selectbox("Nome completo", [""]+list(GESTORES.keys()), key="gestor_carteira")
    g_info = GESTORES.get(gestor, {})
    rb2,rb3 = st.columns(2)
    with rb2: autofill_box("E-mail", g_info.get("email",""))
    with rb3: autofill_box("Telefone", g_info.get("tel",""))

    _s,_b = st.columns([6,2])
    with _b:
        if st.button("Próximo: Endereços →", key="_prox_0", use_container_width=True):
            st.session_state["_click_tab"]=1; st.rerun()

# ─────────────────────────────────────────────
# 02 · ENDEREÇOS
# ─────────────────────────────────────────────
with tabs[1]:
    sec("Bloco 02","Endereços","Endereço fiscal e pontos de entrega.")
    sub("Endereço Fiscal")
    e1,e2,e3 = st.columns([5,1,2])
    with e1: st.text_input("Logradouro", key="end_fiscal_logradouro")
    with e2: st.text_input("Nº", key="end_fiscal_numero")
    with e3: st.text_input("Complemento", key="end_fiscal_complemento")
    e4,e5,e6,e7 = st.columns([3,3,1,2])
    with e4: st.text_input("Bairro", key="end_fiscal_bairro")
    with e5: st.text_input("Município", key="end_fiscal_municipio")
    with e6: st.text_input("UF", max_chars=2, key="end_fiscal_estado")
    with e7: st.text_input("CEP", placeholder="00000-000", key="end_fiscal_cep")

    sub("Endereços de Entrega")
    h1,h2,h3,h4 = st.columns([1,2,3,4])
    col_head("Opção","Identificação","Município / Estado","Observações",cols=[h1,h2,h3,h4])
    for n in range(1,5):
        ec1,ec2,ec3,ec4 = st.columns([1,2,3,4])
        with ec1: st.markdown(f"<div style='padding-top:10px;font-family:IBM Plex Mono,monospace;font-size:.75rem;color:#5a5450'>0{n}</div>",unsafe_allow_html=True)
        with ec2: st.text_input("",key=f"end_entrega_{n}_id",label_visibility="collapsed",placeholder="Nome da unidade")
        with ec3: st.text_input("",key=f"end_entrega_{n}_municipio_estado",label_visibility="collapsed",placeholder="Cidade / UF")
        with ec4: st.text_input("",key=f"end_entrega_{n}_obs",label_visibility="collapsed",placeholder="—")

    _s,_b = st.columns([6,2])
    with _b:
        if st.button("Próximo: Contatos →",key="_prox_1",use_container_width=True):
            st.session_state["_click_tab"]=2; st.rerun()

# ─────────────────────────────────────────────
# 03 · CONTATOS
# ─────────────────────────────────────────────
with tabs[2]:
    sec("Bloco 03","Contatos","Dois contatos por área: financeiro, técnico e logística.")
    contatos = {}
    for g_label, prefix in [("Financeiro","fin"),("Responsável Técnico","tec"),("Logística","log")]:
        sub(g_label)
        h1,h2,h3,h4,h5 = st.columns([1,3,2,2,3])
        col_head("","Nome","Cargo","Tel / Whatsapp","E-mail",cols=[h1,h2,h3,h4,h5])
        for idx in (1,2):
            ic1,ic2,ic3,ic4,ic5 = st.columns([1,3,2,2,3])
            with ic1: st.markdown(f"<div style='padding-top:10px;font-family:IBM Plex Mono,monospace;font-size:.7rem;color:#5a5450'>{idx}°</div>",unsafe_allow_html=True)
            with ic2: contatos[f"{prefix}{idx}_nome"]  = st.text_input("",key=f"{prefix}{idx}_nome", label_visibility="collapsed",placeholder="—")
            with ic3: contatos[f"{prefix}{idx}_cargo"] = st.text_input("",key=f"{prefix}{idx}_cargo",label_visibility="collapsed",placeholder="—")
            with ic4: contatos[f"{prefix}{idx}_tel"]   = st.text_input("",key=f"{prefix}{idx}_tel",  label_visibility="collapsed",placeholder="—")
            with ic5: contatos[f"{prefix}{idx}_email"] = st.text_input("",key=f"{prefix}{idx}_email",label_visibility="collapsed",placeholder="—")

    _s,_b = st.columns([6,2])
    with _b:
        if st.button("Próximo: Comercial →",key="_prox_2",use_container_width=True):
            st.session_state["_click_tab"]=3; st.rerun()

# ─────────────────────────────────────────────
# 04 · COMERCIAL
# ─────────────────────────────────────────────
with tabs[3]:
    sec("Bloco 04","Condições Comerciais e Financeiras","Pagamento e política de bonificação.")
    c1,c2 = st.columns(2)
    with c1: st.text_input("Condição de Pagamento",placeholder="ex: 30/60/90 DDL",key="condicao_pagamento")
    with c2: sim_nao("Política de Bonificação","politica_bonificacao")
    st.text_area("Condições Comerciais Especiais",height=72,key="condicoes_especiais",placeholder="—")
    _s,_b = st.columns([6,2])
    with _b:
        if st.button("Próximo: Regras →",key="_prox_3",use_container_width=True):
            st.session_state["_click_tab"]=4; st.rerun()

# ─────────────────────────────────────────────
# 05 · REGRAS
# ─────────────────────────────────────────────
with tabs[4]:
    sec("Bloco 05","Regras de Faturamento e Logística","Requisitos operacionais para emissão de nota e entrega.")

    sub("Faturamento")
    f1,f2,f3 = st.columns([3,1,3])
    with f1:
        st.selectbox("Data limite para faturamento",[
            "","Até o último dia útil","Sem restrição","Até o dia 25","Personalizado",
        ],key="fat_data_opcao")
        if st.session_state.get("fat_data_opcao")=="Personalizado":
            st.text_input("Especifique:",key="fat_data_custom",placeholder="ex: dia 20")
    with f2: sim_nao("Exige PO?","fat_exige_po")
    with f3: st.text_input("Prazo de Envio para NF",placeholder="ex: 48h antes",key="fat_prazo_nf")
    st.text_area("Observações de Faturamento",height=72,key="fat_observacoes",placeholder="—")

    sub("Logística")
    l1,l2,l3 = st.columns([2,1,2])
    with l1: st.selectbox("Tipo de entrega",["","CIF","FOB"],key="log_tipo_entrega")
    with l2: sim_nao("Agendamento?","log_agendamento")
    with l3: st.text_input("Prazo mínimo agendamento",placeholder="ex: 48h",key="log_prazo_agendamento")

    l4,l5 = st.columns(2)
    with l4:
        st.selectbox("Dias de recebimento",[
            "","Todos os dias","Todos os dias úteis","Personalizado",
        ],key="log_dias_opcao")
        if st.session_state.get("log_dias_opcao")=="Personalizado":
            st.multiselect("Selecione os dias:",
                ["Segunda","Terça","Quarta","Quinta","Sexta","Sábado","Domingo"],
                key="log_dias_sel")
    with l5:
        st.selectbox("Horário de recebimento",["","Horário comercial","Personalizado"],key="log_horario_opcao")
        if st.session_state.get("log_horario_opcao")=="Personalizado":
            h1,h2 = st.columns(2)
            with h1: st.text_input("De:",key="log_horario_de",placeholder="07:00")
            with h2: st.text_input("Até:",key="log_horario_ate",placeholder="17:00")
    st.text_area("Regras de acesso / EPI / Observações",height=72,key="log_acesso_obs",placeholder="—")

    _s,_b = st.columns([6,2])
    with _b:
        if st.button("Próximo: Licenças →",key="_prox_4",use_container_width=True):
            st.session_state["_click_tab"]=5; st.rerun()

# ─────────────────────────────────────────────
# 06 · LICENÇAS
# ─────────────────────────────────────────────
with tabs[5]:
    sec("Bloco 06","Licenças e Documentação do Cliente","Status de validade das licenças regulatórias do cliente.")
    lic_data = {}
    hd = st.columns([3,2,2,4])
    col_head("Documento","Status","Validade","Observações",cols=hd)
    for nome_lic, prefix in [
        ("Licença Ambiental","lic_ambiental"),
        ("Licença de Operação","lic_operacao"),
        ("Registro MAPA","lic_mapa"),
        ("Alvará Municipal","lic_alvara"),
        ("Certificação ISO","lic_iso"),
    ]:
        c1,c2,c3,c4 = st.columns([3,2,2,4])
        with c1: st.markdown(f"<div style='padding-top:10px;font-size:.85rem;color:#a09890'>{nome_lic}</div>",unsafe_allow_html=True)
        with c2: lic_data[f"{prefix}_status"]   = s_lic("",f"{prefix}_status",lv="collapsed")
        with c3: lic_data[f"{prefix}_validade"] = st.date_input("",value=None,key=f"{prefix}_validade",label_visibility="collapsed",format="DD/MM/YYYY")
        with c4: lic_data[f"{prefix}_obs"]      = st.text_input("",key=f"{prefix}_obs",label_visibility="collapsed",placeholder="—")

    _s,_b = st.columns([6,2])
    with _b:
        if st.button("Próximo: Estratégico →",key="_prox_5",use_container_width=True):
            st.session_state["_click_tab"]=6; st.rerun()

# ─────────────────────────────────────────────
# 07 · ESTRATÉGICO
# ─────────────────────────────────────────────
with tabs[6]:
    sec("Bloco 07","Complexidade Operacional","Avaliação dos critérios operacionais do cliente.")
    hd = st.columns([4,1,4])
    col_head("Critério","Nível (1–3)","Observação",cols=hd)
    compl_data = {}
    for nome_crit, prefix in [
        ("Exigências logísticas","compl_log"),
        ("Burocracia de faturamento","compl_fat"),
        ("Dificuldade de acesso","compl_acesso"),
        ("Nível de exigência operacional","compl_operacional"),
    ]:
        cr1,cr2,cr3 = st.columns([4,1,4])
        with cr1: st.markdown(f"<div style='padding-top:10px;font-size:.85rem;color:#a09890'>{nome_crit}</div>",unsafe_allow_html=True)
        with cr2: compl_data[f"{prefix}_nivel"] = nivel("",f"{prefix}_nivel",lv="collapsed")
        with cr3: compl_data[f"{prefix}_obs"]   = st.text_input("",key=f"{prefix}_obs",label_visibility="collapsed",placeholder="—")

    _s,_b = st.columns([6,2])
    with _b:
        if st.button("Próximo: Crédito →",key="_prox_6",use_container_width=True):
            st.session_state["_click_tab"]=7; st.rerun()

# ─────────────────────────────────────────────
# 08 · CRÉDITO + CONTROLE
# ─────────────────────────────────────────────
with tabs[7]:
    cred_data = {}

    def render_docs(prefix, bloco, titulo, docs):
        sec(f"Bloco {bloco}", titulo)
        hd = st.columns([4,3,2,3])
        col_head("Documento","Arquivo","Data de entrega","Observações",cols=hd)
        for n, nome_doc in enumerate(docs, start=1):
            c1,c2,c3,c4 = st.columns([4,3,2,3])
            with c1:
                st.markdown(f"<div style='padding-top:10px;font-size:.82rem;color:#a09890'>{nome_doc}</div>",unsafe_allow_html=True)
            with c2:
                st.file_uploader("",key=f"{prefix}_doc{n}_file",
                                 label_visibility="collapsed",
                                 type=["pdf","jpg","jpeg","png","docx"])
            with c3:
                cred_data[f"{prefix}_doc{n}_data"] = st.date_input(
                    "",value=None,key=f"{prefix}_doc{n}_data",
                    label_visibility="collapsed",format="DD/MM/YYYY")
            with c4:
                cred_data[f"{prefix}_doc{n}_obs"] = st.text_input(
                    "",key=f"{prefix}_doc{n}_obs",
                    label_visibility="collapsed",placeholder="—")

    render_docs("cred_ltda","08a","Análise de Crédito — Empresas LTDA", DOCS_LTDA)
    render_docs("cred_coop","08b","Análise de Crédito — Cooperativas / S.A. / Usinas", DOCS_COOP)

    sec("Controle","Controle Interno")
    ci1,ci2,ci3,ci4 = st.columns(4)
    with ci1:
        st.text_input("Cadastro realizado por",key="ctrl_cadastrado_por")
        st.date_input("Data do cadastro",value=date.today(),key="ctrl_data_cadastro",format="DD/MM/YYYY")
    with ci2:
        st.text_input("Responsável pela atualização",key="ctrl_responsavel_atualizacao")
        st.date_input("Última atualização",value=date.today(),key="ctrl_ultima_atualizacao",format="DD/MM/YYYY")
    with ci3:
        st.text_input("Revisado por",key="ctrl_revisadopor")
    with ci4:
        st.text_input("Aprovado por",key="ctrl_aprovadopor")
        st.selectbox("Status do cadastro",["","Em preenchimento","Completo","Em revisão","Aprovado"],key="ctrl_status")

    _s,_b = st.columns([6,2])
    with _b:
        if st.button("Próximo: Produtos →",key="_prox_7",use_container_width=True):
            st.session_state["_click_tab"]=8; st.rerun()

# ─────────────────────────────────────────────
# 09 · PRODUTOS DO CLIENTE
# ─────────────────────────────────────────────
with tabs[8]:
    sec("Bloco 09","Produtos do Cliente",
        "Relação de produtos, fichas técnicas ATLAS e shelf-life mínimo.")
    num_prod = st.number_input(
        "Quantidade de produtos (Máximo 14)",
        min_value=0, max_value=14, step=1, value=0, key="num_produtos")

    if int(num_prod) > 0:
        st.markdown("<div style='margin-top:8px'></div>",unsafe_allow_html=True)
        hp0,hp1,hp2,hp3 = st.columns([0.4,3,3,3])
        col_head("","Nome / Cód. do Produto","Ficha Técnica ATLAS","Shelf-life mínimo para entrega",
                 cols=[hp0,hp1,hp2,hp3])
        for i in range(1, int(num_prod)+1):
            pc0,pc1,pc2,pc3 = st.columns([0.4,3,3,3])
            with pc0:
                st.markdown(f"<div style='padding-top:10px;font-family:IBM Plex Mono,monospace;font-size:.75rem;color:#5a5450'>{i:02d}</div>",unsafe_allow_html=True)
            with pc1:
                st.text_input("",key=f"produtos_do_c_{i}",label_visibility="collapsed",placeholder="—")
            with pc2:
                st.file_uploader("",key=f"atlas_file_{i}",label_visibility="collapsed",
                                 type=["pdf","jpg","jpeg","png","docx","xlsx"])
            with pc3:
                st.text_input("",key=f"SFMINdoprod_{i}",label_visibility="collapsed",placeholder="—")

    st.markdown("<div style='margin-top:32px;border-top:1px solid rgba(255,255,255,0.06);padding-top:28px'></div>",
                unsafe_allow_html=True)
    _s,_b = st.columns([6,2])
    with _b:
        if st.button("Gerar Ficha",key="_gerar_btn",type="primary",use_container_width=True):
            st.session_state["_gerar"]=True

# ──────────────────────────────────────────────────────────────────────────────
# Download display (outside tabs)
# ──────────────────────────────────────────────────────────────────────────────
if "arquivo_gerado" in st.session_state:
    is_zip = st.session_state.get("_is_zip", False)
    st.markdown("""
    <div style="text-align:center;margin-top:20px">
      <div style="font-size:.72rem;color:#10e68d;letter-spacing:.06em;
                  text-transform:uppercase;margin-bottom:12px">✓ Ficha gerada com sucesso</div>
    </div>""", unsafe_allow_html=True)
    if is_zip:
        st.markdown("""<div style="text-align:center;font-size:.68rem;color:#766f6b;margin-bottom:8px">
        Extraia o arquivo ZIP e abra o Excel na mesma pasta para que os links dos documentos funcionem.
        </div>""", unsafe_allow_html=True)
    col_l,col_dl,col_r = st.columns([2,3,2])
    with col_dl:
        label = "↓  Baixar Pasta Compactada (ZIP)" if is_zip else "↓  Baixar Excel"
        mime  = "application/zip" if is_zip else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        st.download_button(label,
            data=st.session_state["arquivo_gerado"],
            file_name=st.session_state.get("_nome_arquivo","PRISMA.xlsx"),
            mime=mime, use_container_width=True)

st.markdown("<div style='margin-bottom:60px'></div>",unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# LÓGICA — Gerar Ficha
# ══════════════════════════════════════════════════════════════════════════════
if st.session_state.get("_gerar"):
    st.session_state["_gerar"] = False
    razao = st.session_state.get("razao_social","").strip()
    if not razao:
        st.error("Preencha a Razão Social (aba Cadastro) antes de gerar a ficha.")
    else:
        # ── Gestor ──
        gestor_name = st.session_state.get("gestor_carteira","")
        gestor_info = GESTORES.get(gestor_name,{})

        # ── Faturamento: data limite ──
        fat_opcao  = st.session_state.get("fat_data_opcao","")
        fat_custom = st.session_state.get("fat_data_custom","")
        fat_data_limite = fat_custom if fat_opcao=="Personalizado" else fat_opcao

        # ── Logística: dias ──
        dias_op  = st.session_state.get("log_dias_opcao","")
        dias_sel = st.session_state.get("log_dias_sel",[])
        log_dias = ", ".join(dias_sel) if (dias_op=="Personalizado" and dias_sel) else dias_op

        # ── Logística: horário ──
        hora_op  = st.session_state.get("log_horario_opcao","")
        hora_de  = st.session_state.get("log_horario_de","")
        hora_ate = st.session_state.get("log_horario_ate","")
        log_horario = (f"De {hora_de} às {hora_ate}" if (hora_de or hora_ate) else hora_op) \
                      if hora_op=="Personalizado" else hora_op

        # ── Número de produtos ──
        n_prod = int(st.session_state.get("num_produtos",0))

        log_obs = st.session_state.get("log_acesso_obs","")

        # ── TEXT DATA ──
        text_data = {
            "razao_social":               razao,
            "nome_fantasia":              st.session_state.get("nome_fantasia",""),
            "cnpj":                       st.session_state.get("cnpj",""),
            "inscricao_estadual":         st.session_state.get("inscricao_estadual",""),
            "inscricao_municipal":        st.session_state.get("inscricao_municipal",""),
            "cnae":                       st.session_state.get("cnae",""),
            "data_abertura":              fmt_date(st.session_state.get("data_abertura")),
            "ctrl_responsavel_biotrop":           gestor_name,
            "ctrl_responsavel_biotrop-email":     gestor_info.get("email",""),
            "ctrl_responsavel_biotrop-telefone":  gestor_info.get("tel",""),
            "grupo_economico":"", "data_inicio_relacionamento":"",
            "end_fiscal_logradouro":  st.session_state.get("end_fiscal_logradouro",""),
            "end_fiscal_numero":      st.session_state.get("end_fiscal_numero",""),
            "end_fiscal_complemento": st.session_state.get("end_fiscal_complemento",""),
            "end_fiscal_bairro":      st.session_state.get("end_fiscal_bairro",""),
            "end_fiscal_municipio":   st.session_state.get("end_fiscal_municipio",""),
            "end_fiscal_estado":      st.session_state.get("end_fiscal_estado",""),
            "end_fiscal_cep":         st.session_state.get("end_fiscal_cep",""),
            **{k: st.session_state.get(k,"") for k in contatos},
            "com1_nome":"","com1_cargo":"","com1_tel":"","com1_email":"",
            "com2_nome":"","com2_cargo":"","com2_tel":"","com2_email":"",
            "tipo_cliente":"","culturas_principais":"","area_total_ha":"",
            "regiao_atuacao":"","produtos_utilizados":"",
            "condicao_pagamento":   st.session_state.get("condicao_pagamento",""),
            "politica_bonificacao": st.session_state.get("politica_bonificacao",""),
            "condicoes_especiais":  st.session_state.get("condicoes_especiais",""),
            "prazo_medio_dias":"","limite_credito":"","forma_pagamento":"",
            "fat_data_limite":    fat_data_limite,
            "fat_exige_po":       st.session_state.get("fat_exige_po",""),
            "fat_prazo_nf":       st.session_state.get("fat_prazo_nf",""),
            "fat_observacoes":    st.session_state.get("fat_observacoes",""),
            "fat_exige_contrato":"","fat_conferencia_nf":"","fat_shelf_life":"",
            "log_tipo_entrega":        st.session_state.get("log_tipo_entrega",""),
            "log_agendamento":         st.session_state.get("log_agendamento",""),
            "log_prazo_agendamento":   st.session_state.get("log_prazo_agendamento",""),
            "log_dias_recebimento":    log_dias,
            "log_horario_recebimento": log_horario,
            "log_regras_acesso":       log_obs,
            "log_observacoes":         log_obs,
            "log_aviso_entrega":"","log_antecedencia_aviso":"",
            "log_nf_antecipada":"","log_antecedencia_nf":"",
            "log_romaneio":"","log_restricao_transportadora":"",
            **{k: fmt_date(v) if isinstance(v,(date,datetime)) else (v or "")
               for k,v in lic_data.items()},
            **{k: (v or "") for k,v in compl_data.items()},
            "compl_classificacao":"","compl_justificativa":"",
            "estrat_concorrentes":"","estrat_potencial_volume":"",
            "estrat_participacao":"","estrat_historico":"",
            "estrat_risco":"","estrat_observacoes":"",
            **{k: fmt_date(v) if isinstance(v,(date,datetime)) else (v or "")
               for k,v in cred_data.items()},
            "ctrl_cadastrado_por":          st.session_state.get("ctrl_cadastrado_por",""),
            "ctrl_data_cadastro":           fmt_date(st.session_state.get("ctrl_data_cadastro")),
            "ctrl_ultima_atualizacao":      fmt_date(st.session_state.get("ctrl_ultima_atualizacao")),
            "ctrl_responsavel_atualizacao": st.session_state.get("ctrl_responsavel_atualizacao",""),
            "ctrl_status":                  st.session_state.get("ctrl_status",""),
            "ctl_revisadopor":              st.session_state.get("ctrl_revisadopor",""),
            "ctl_aprovadopor":              st.session_state.get("ctrl_aprovadopor",""),
        }

        # Endereços de entrega
        for n in range(1,5):
            for campo in ("id","municipio_estado","obs"):
                text_data[f"end_entrega_{n}_{campo}"] = st.session_state.get(f"end_entrega_{n}_{campo}","")

        # Produtos — texto (nome e sfmin)
        for i in range(1,15):
            if i <= n_prod:
                text_data[f"produtos_do_c_{i}"] = st.session_state.get(f"produtos_do_c_{i}","")
                text_data[f"SFMINdoprod_{i}"]   = st.session_state.get(f"SFMINdoprod_{i}","")
            else:
                text_data[f"produtos_do_c_{i}"] = ""
                text_data[f"SFMINdoprod_{i}"]   = ""

        # ── FILE DATA ──
        file_data = {}

        # ATLAS por produto
        for i in range(1,15):
            fobj = st.session_state.get(f"atlas_file_{i}")
            if fobj and i <= n_prod:
                prod_name = (st.session_state.get(f"produtos_do_c_{i}","") or f"Produto_{i}").strip()
                safe_prod = _safe_filename(prod_name)
                ext = fobj.name.rsplit(".",1)[-1].lower() if "." in fobj.name else "pdf"
                saved = f"Ficha ATLAS - {safe_prod}.{ext}"
                fobj.seek(0)
                file_data[f"ATLAS-do-prod-{i}"] = {
                    "saved_name": saved,
                    "display":    f"Ficha ATLAS - {prod_name}",
                    "bytes":      fobj.read(),
                }
            else:
                # i > n_prod → "-" via text_data; within range no file → Pendente via file_data
                if i <= n_prod:
                    file_data[f"ATLAS-do-prod-{i}"] = None  # → Pendente
                else:
                    text_data[f"ATLAS-do-prod-{i}"] = ""    # → "-"

        # Documentos de crédito
        for prefix, docs in [("cred_ltda",DOCS_LTDA),("cred_coop",DOCS_COOP)]:
            for n, nome_doc in enumerate(docs, start=1):
                fobj = st.session_state.get(f"{prefix}_doc{n}_file")
                if fobj:
                    safe_doc = _safe_filename(nome_doc)
                    ext = fobj.name.rsplit(".",1)[-1].lower() if "." in fobj.name else "pdf"
                    saved = f"{safe_doc}.{ext}"
                    fobj.seek(0)
                    file_data[f"{prefix}_doc{n}_status"] = {
                        "saved_name": saved,
                        "display":    fobj.name,
                        "bytes":      fobj.read(),
                    }
                else:
                    file_data[f"{prefix}_doc{n}_status"] = None  # → Pendente

        # ── GENERATE ──
        try:
            excel_bytes, files_to_include = fill_template(text_data, file_data)
            safe_razao = _safe_filename(razao)

            if files_to_include:
                output = build_zip(excel_bytes, files_to_include, razao)
                fname  = f"PRISMA_{safe_razao}.zip"
                is_zip = True
            else:
                output = excel_bytes
                fname  = f"PRISMA_{safe_razao}.xlsx"
                is_zip = False

            st.session_state["arquivo_gerado"] = output
            st.session_state["_nome_arquivo"]  = fname
            st.session_state["_is_zip"]        = is_zip
            st.rerun()
        except FileNotFoundError:
            st.error("Template não encontrado em `template/Prisma - Template.xlsx`.")
        except Exception as e:
            st.error(f"Erro ao gerar a ficha: {e}")
